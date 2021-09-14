import bs4
import urllib.request as req
from openpyxl.styles import colors
from openpyxl.styles.colors import Color
from openpyxl.utils.cell import cols_from_range
import requests
import os
from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import Alignment
from time import sleep
from tqdm import tqdm
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
import rarity
import shelve

# # 建立試算表
# wb=Workbook()
# ws=wb.active

seed='3109254173'
event='2021-08-30_607'
lang='tw'

seedvalue=shelve.open('lastseed')
seedvalue['seed']=seed
seedvalue['event']=event
seedvalue['lang']=lang
seedvalue.close()
# # 取得貓咪列表
# rare=rarity.get_rare()
# supa=rarity.get_supa()
# uber=rarity.get_uber()
# legend=rarity.get_legend()
# blue=rarity.get_blue()

# # 取得網頁內容
# url=f"https://bc.godfat.org/?seed={seed}&event={event}&lang={lang}"
# User_Agent={'user_agent':"Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/93.0.4577.63 Safari/537.36"}

# root=requests.get(url,headers=User_Agent)
# htmlfile=bs4.BeautifulSoup(root.text,"lxml")

# # 取得活動時間和內容
# event_date=[]
# event_pool=[]
# events=htmlfile.find("optgroup",label="Upcoming:").find_all("option")
# for event in events:
#     event_date.append(event.string.strip()[0:23])
#     event_pool.append(event.string.strip()[25:])

# title_length=4*len(event_pool)

# # 插入活動時間和內容，順便調整欄位
# ws.append(event_date)
# ws.append(event_pool)
# for i in range(1,title_length+4):
#     ws[f'{get_column_letter(i)}2'].font=Font(size=14,name='Noto Sans CJK TC')
# for i in range(1,title_length+4):
#     ws[f'{get_column_letter(i)}1'].font=Font(size=12,name='Liberation Sans')
# for add in range(2,title_length,4):
#     ws.insert_cols(add)
#     ws.insert_cols(add)
#     ws.insert_cols(add)
# # 設定欄寬列高
# ws.row_dimensions[2].height=80
# ws.row_dimensions[1].height=37
# for index in range(1,title_length+2,2):
#     ws.column_dimensions[str(get_column_letter(index))].width=20
# for index in range(2,title_length+3,2):
#     ws.column_dimensions[str(get_column_letter(index))].width=25


# ws.column_dimensions['A'].width=5
# ws.freeze_panes='B1'
# for number in range(3,104):
#     ws.row_dimensions[number].height=20

# # 設定文字自適應
# for row in ws.iter_rows(1,3):  
#     for cell in row:      
#         cell.alignment = Alignment(wrap_text=True) 

# # 插入AB標題
# for ab in range(1,title_length+2,4):
#         ws[f'{get_column_letter(ab)}3'].value='A'
#         ws[f'{get_column_letter(ab+1)}3'].value='Guaranteed'
#         ws[f'{get_column_letter(ab+2)}3'].value='B'
#         ws[f'{get_column_letter(ab+3)}3'].value='Guaranteed'

# # 取得所有活動碼
# events_header=[]
# events_headers=htmlfile.find("optgroup",label="Upcoming:").find_all("option")
# for event in events_headers:
#     events_header.append(event["value"])


# # 建立進度條
# times = 0
# progress = tqdm(total=len(events_header))

# # 設定樣式
# style_cat_name=Font(size=16,name='Noto Sans CJK TC')
# style_gran_name=Font(size=12,name='Noto Sans CJK TC',color='00800080')
# style_num=Font(size=14,name='Liberation Sans')

# style_rare_fill=PatternFill(start_color="00CCFFCC", end_color="00CCFFCC", fill_type="solid")
# style_supa_fill=PatternFill(start_color="00FF9900", end_color="00FF9900",fill_type="solid")
# style_uber_fill=PatternFill(start_color="00FF0000", end_color="00FF0000",fill_type="solid")
# style_legend_fill=PatternFill(start_color="00CC99FF", end_color="00CC99FF",fill_type="solid")
# style_blue_fill=PatternFill(start_color="0000CCFF", end_color="0000CCFF",fill_type="solid")
# style_change_row_fill=PatternFill(start_color="00969696", end_color="00969696",fill_type="solid")

# #取得各活動的貓咪序列
# row=1
# for event_header in events_header:#取得某活動的網址
#     url=f"https://bc.godfat.org/?seed={seed}&event={event_header}&lang={lang}"
#     User_Agent={'user_agent':"Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/93.0.4577.63 Safari/537.36"}

#     root=requests.get(url,headers=User_Agent)
#     htmlfile=bs4.BeautifulSoup(root.text,"lxml")

#     # 取得貓咪AB序列
#     numsA=range(1,101)
#     roolA_name=[]
#     numsB=range(1,101)
#     roolB_name=[]
#     realityA=[]
#     realityB=[]
#     roolAG_name=[]
#     roolBG_name=[]
#     for numA in numsA:
#         roolA=htmlfile.find("td", onclick=f"pick('{numA}A')")
#         roolAR=htmlfile.find("td", onclick=f"pick('{numA}AR')")
#         roolAG=htmlfile.find("td", onclick=f"pick('{numA}AG')")
#         if roolAR==None:
#             roolA_name.append(roolA.a.string)
#             if roolAG!=None:
#                 roolAG_name.append(roolAG.a.string)
#         else:
#             roolA_name.append(roolA.a.string+"→"+roolAR.a.string)
#             if roolAG!=None:
#                 roolAG_name.append(roolAG.a.string)
#     for numB in numsB:
#         roolB=htmlfile.find("td", onclick=f"pick('{numB}B')")
#         roolBR=htmlfile.find("td", onclick=f"pick('{numB}BR')")
#         roolBG=htmlfile.find("td", onclick=f"pick('{numB}BG')")
#         if roolBR==None:
#             roolB_name.append(roolB.a.string)
#             if roolBG!=None:
#                 roolBG_name.append(roolBG.a.string)
#         else:
#             roolB_name.append(roolB.a.string+"→"+roolBR.a.string)
#             if roolBG!=None:
#                 roolBG_name.append(roolBG.a.string)

# # 填入貓咪序列
#     for col in range(1,101):
#         ws[f'{get_column_letter(row)}{col+3}'].value=roolA_name[col-1]
#         if roolAG_name!=[]:
#             for i in range(1,91):
#                 ws[f'{get_column_letter(row+1)}{i+3}'].value=roolAG_name[i-1]
#                 ws[f'{get_column_letter(row+1)}{i+3}'].font=style_gran_name
#         if roolA_name[col-1] in rare:
#             ws[f'{get_column_letter(row)}{col+3}'].font=style_cat_name
#             ws[f'{get_column_letter(row)}{col+3}'].fill=style_rare_fill
#         elif roolA_name[col-1] in supa:
#             ws[f'{get_column_letter(row)}{col+3}'].font=style_cat_name
#             ws[f'{get_column_letter(row)}{col+3}'].fill=style_supa_fill
#         elif roolA_name[col-1] in uber:
#             ws[f'{get_column_letter(row)}{col+3}'].font=style_cat_name
#             ws[f'{get_column_letter(row)}{col+3}'].fill=style_uber_fill
#         elif roolA_name[col-1] in legend:
#             ws[f'{get_column_letter(row)}{col+3}'].font=style_cat_name
#             ws[f'{get_column_letter(row)}{col+3}'].fill=style_legend_fill
#         elif roolA_name[col-1] in blue:
#             ws[f'{get_column_letter(row)}{col+3}'].font=style_cat_name
#             ws[f'{get_column_letter(row)}{col+3}'].fill=style_blue_fill
#         else:
#             ws[f'{get_column_letter(row)}{col+3}'].font=style_cat_name
#             ws[f'{get_column_letter(row)}{col+3}'].fill=style_change_row_fill
#     for col in range(1,101):
#         ws[f'{get_column_letter(row+2)}{col+3}'].value=roolB_name[col-1]
#         if roolBG_name!=[]:
#             for i in range(1,90):
#                 ws[f'{get_column_letter(row+3)}{i+3}'].value=roolBG_name[i-1]
#                 ws[f'{get_column_letter(row+3)}{i+3}'].font=style_gran_name
#         if roolB_name[col-1] in rare:
#             ws[f'{get_column_letter(row+2)}{col+3}'].font=style_cat_name
#             ws[f'{get_column_letter(row+2)}{col+3}'].fill=style_rare_fill
#         elif roolB_name[col-1] in supa:
#             ws[f'{get_column_letter(row+2)}{col+3}'].font=style_cat_name
#             ws[f'{get_column_letter(row+2)}{col+3}'].fill=style_supa_fill
#         elif roolB_name[col-1] in uber:
#             ws[f'{get_column_letter(row+2)}{col+3}'].font=style_cat_name
#             ws[f'{get_column_letter(row+2)}{col+3}'].fill=style_uber_fill
#         elif roolB_name[col-1] in legend:
#             ws[f'{get_column_letter(row+2)}{col+3}'].font=style_cat_name
#             ws[f'{get_column_letter(row+2)}{col+3}'].fill=style_legend_fill
#         elif roolB_name[col-1] in blue:
#             ws[f'{get_column_letter(row+2)}{col+3}'].font=style_cat_name
#             ws[f'{get_column_letter(row+2)}{col+3}'].fill=style_blue_fill
#         else:
#             ws[f'{get_column_letter(row+2)}{col+3}'].font=style_cat_name
#             ws[f'{get_column_letter(row+2)}{col+3}'].fill=style_change_row_fill
#     row+=4
#     sleep(0.01)
#     progress.update(1)

# # 插入序列號碼
# ws.insert_cols(1)
# for Number in range(1,101):
#     ws[f'A{Number+3}'].value=Number
#     ws[f'A{Number+3}'].font=style_num

# # 存檔
# wb.save('seedlists.xlsx')