import os
from openpyxl import Workbook, load_workbook
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import Font
from openpyxl.styles import PatternFill

wb=openpyxl.load_workbook('seedlists.xlsx')
ws=wb.active

ws.column_dimensions["A"].width=5
wb.save('seedlists.xlsx')